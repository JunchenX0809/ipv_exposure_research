#!/usr/bin/env python3
"""Build the Zimbabwe 2017 five-column outcome codebook."""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import pyreadstat
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from utils.covariates import classify_variable_type
from utils.docx_export import export_minimal_codebook_docx
from utils.outcomes import OUTCOME_CODEBOOK_COLUMNS, build_outcome_codebook_df, outcome_codebook_to_tsv

RAW = ROOT / "data" / "raw" / "Zimbabwe Stata"
PUD = RAW / "ZIMBABWE_VACS_2017_PUD.dta"
RESPONDENT_CODEBOOK = RAW / "ZIMBABWE_VACS_2017_Respondent_Codebook.xlsx"
OUT_DIR = ROOT / "data" / "processed" / "outcome"
OUT_DOCX = OUT_DIR / "jcx_zimbabweOutcomeCodebook.docx"
OUT_TSV = OUT_DIR / "jcx_zimbabweOutcomeCodebook.tsv"


def _parse_response_cell(value: object) -> tuple[str, str] | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower().startswith(("total", "frequency missing")):
        return None
    if "=" in text:
        code, label = text.split("=", 1)
        code = code.strip()
        label = label.strip()
        if code and code.replace(".", "", 1).isdigit() and label:
            return code, label
    if text.isdigit():
        return text, text
    return None


def _read_respondent_codebook() -> dict[str, dict[str, str]]:
    wb = load_workbook(RESPONDENT_CODEBOOK, read_only=True, data_only=True)
    out: dict[str, dict[str, str]] = {}
    try:
        for ws in wb.worksheets:
            values = list(ws.iter_rows(values_only=True))
            for i, row in enumerate(values):
                first = "" if not row or row[0] is None else str(row[0]).strip()
                second = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
                if not first or second.lower() != "skip":
                    continue
                var = first.lower()
                question = ""
                if i > 0 and values[i - 1] and values[i - 1][0] is not None:
                    question = str(values[i - 1][0]).strip()
                responses: list[tuple[str, str]] = []
                j = i + 1
                while j < len(values):
                    next_first = "" if not values[j] or values[j][0] is None else str(values[j][0]).strip()
                    next_second = "" if len(values[j]) < 2 or values[j][1] is None else str(values[j][1]).strip()
                    if not next_first or next_second.lower() == "skip":
                        break
                    parsed = _parse_response_cell(next_first)
                    if parsed:
                        responses.append(parsed)
                    j += 1
                out[var] = {
                    "question": question,
                    "format": ", ".join(f"{code}-{label}" for code, label in responses),
                }
    finally:
        wb.close()
    return out


def _fmt(codebook: dict[str, dict[str, str]], vars_: list[str]) -> str:
    parts = []
    for var in vars_:
        entry = codebook.get(var.lower())
        if not entry:
            continue
        fmt = entry.get("format", "")
        parts.append(f"{var}: {fmt}" if fmt else f"{var}:")
    return "; ".join(parts)


def _questions(codebook: dict[str, dict[str, str]], vars_: list[str]) -> str:
    texts: list[str] = []
    for var in vars_:
        override = QUESTION_OVERRIDES.get(var.lower())
        if override:
            texts.append(override)
            continue
        entry = codebook.get(var.lower())
        if entry and entry.get("question"):
            texts.append(entry["question"])
    return " || ".join(dict.fromkeys(texts))


QUESTION_OVERRIDES = {
    "q116b": (
        "116B. Has a person within your age range ever: B. choked, smothered, "
        "tried to drown you, or burned you intentionally?"
    ),
}


MALE_PERPETRATION_QUESTIONS = {
    "q200a": (
        "200A. Have you ever done any of the following to a current or previous girlfriend, "
        "romantic partner, wife, or casual sex partner: A. punched, kicked, whipped, or beat them?"
    ),
    "q200b": (
        "200B. Have you ever done any of the following to a current or previous girlfriend, "
        "romantic partner, wife, or casual sex partner: B. choked, smothered, tried to drown, "
        "or intentionally burn them?"
    ),
    "q200c": (
        "200C. Have you ever done any of the following to a current or previous girlfriend, "
        "romantic partner, wife, or casual sex partner: C. used or threatened to use a knife, "
        "gun, knobkerrie or other weapon against them?"
    ),
    "q201a": (
        "201A. Have you ever done any of the following to someone who is not a current or previous "
        "girlfriend, romantic partner, wife, or casual sex partner: A. punched, kicked, whipped, "
        "or beat them?"
    ),
    "q201b": (
        "201B. Have you ever done any of the following to someone who is not a current or previous "
        "girlfriend, romantic partner, wife, or casual sex partner: B. choked, smothered, tried "
        "to drown, or intentionally burn them?"
    ),
    "q201c": (
        "201C. Have you ever done any of the following to someone who is not a current or previous "
        "girlfriend, romantic partner, wife, or casual sex partner: C. used or threatened to use "
        "a knife, gun, knobkerrie or other weapon against them?"
    ),
    "q1100": (
        "1100. Have you ever forced a girlfriend/romantic partner, ex-girlfriend/romantic partner, "
        "casual sex partner, wife, or ex-wife to have sex with you when they did not want to?"
    ),
}


def _male_question(var: str) -> str:
    return MALE_PERPETRATION_QUESTIONS.get(var.lower(), "")


def _row(category: str, vars_: list[str] | None = None, *, questions_vars: list[str] | None = None, male_only: bool = False) -> dict[str, object]:
    return {
        "category": category,
        "vars": vars_ or [],
        "questions_vars": questions_vars,
        "male_only": male_only,
    }


OUTCOME_SPEC = [
    _row("Hit etc from parents/caregivers/adult relatives in the last 12 months"),
    _row("Hit with object etc from parents/caregivers/adult relatives in the last 12 months", ["q128a", "q130"]),
    _row("Choked etc from parents/caregivers/adult relatives in the last 12 months", ["q128b", "q130"]),
    _row("Burned etc from parents/caregivers/adult relatives in the last 12 months", ["q128b", "q130"]),
    _row("Threatened with knife/weapon etc from parents/caregivers/adult relatives in the last 12 months", ["q128c", "q130"]),
    _row("Hit etc from public authority figure in the last 12 months", ["q142a", "q144", "q146"], questions_vars=["q142a", "q146"]),
    _row("Choked etc from public authority figure in the last 12 months", ["q142b", "q144", "q146"], questions_vars=["q142b", "q146"]),
    _row("Burned etc from public authority figure in the last 12 months", ["q142b", "q144", "q146"], questions_vars=["q142b", "q146"]),
    _row("Threatened with knife/weapon etc from public authority figure in the last 12 months", ["q142c", "q144", "q146"], questions_vars=["q142c", "q146"]),
    _row("Hit etc from intimate partner in the last 12 months"),
    _row("Hit with object etc from intimate partner in the last 12 months", ["q100a", "q102"]),
    _row("Choked etc from intimate partner in the last 12 months", ["q100b", "q102"]),
    _row("Threatened with knife/weapon etc from intimate partner in the last 12 months", ["q100c", "q102"]),
    _row("Hit etc from peer in the last 12 months"),
    _row("Hit with object etc from peer in the last 12 months", ["q116a", "q118"]),
    _row("Choked etc from peer in the last 12 months", ["q116b", "q118"]),
    _row("Threatened with knife/weapon etc from peer in the last 12 months", ["q116c", "q118"]),
    _row("Hit etc from neighbor in the last 12 months", ["q142a", "q144", "q146"], questions_vars=["q142a", "q146"]),
    _row("Hit with object etc from neighbor in the last 12 months", ["q142a", "q144", "q146"], questions_vars=["q142a", "q146"]),
    _row("Choked etc from neighbor in the last 12 months", ["q142b", "q144", "q146"], questions_vars=["q142b", "q146"]),
    _row("Threatened with knife/weapon etc from neighbor in the last 12 months", ["q142c", "q144", "q146"], questions_vars=["q142c", "q146"]),
    _row("Witnessed parents/caregivers/adult relatives hit etc in the last 12 months", ["q49", "q50"]),
    _row("Hit etc from teacher in the last 12 months", ["q142a", "q144", "q146"], questions_vars=["q142a", "q146"]),
    _row("Offensive names online in the last 12 months"),
    _row("Physically threatened online in the last 12 months"),
    _row("Harassed for sustained period online in the last 12 months"),
    _row("Stalked online in the last 12 months"),
    _row("Purposely embarrassed online in the last 12 months"),
    _row("Not attended school due to safety in the last 12 months"),
    _row("Said not loved etc from parents/caregivers/adult relatives in the last 12 months", ["q300a", "q302"]),
    _row("Wished you were not born or dead etc from parents/caregivers/adult relatives in the last 12 months", ["q300b", "q302"]),
    _row("Ridiculed you etc from parents/caregivers/adult relatives in the last 12 months", ["q300c", "q302"]),
    _row("Threatened to abandon you etc from parents/caregivers/adult relatives in the last 12 months"),
    _row("Shouted at you etc from parents/caregivers/adult relatives in the last 12 months"),
    _row("Take away privileges etc from parents/caregivers/adult relatives in the last 12 months"),
    _row("Insulted you in front of others etc from intimate partner in the last 12 months"),
    _row("Kept you from having your own money etc from intimate partner in the last 12 months"),
    _row("Kept you from talking to friends or family etc from intimate partner in the last 12 months"),
    _row("Demanded to know where you were etc from intimate partner in the last 12 months"),
    _row("Threatened to physically harm you etc from intimate partner in the last 12 months"),
    _row("Made you feel scared from saying mean things etc from peer in the last 12 months"),
    _row("Told lies etc from peer in the last 12 months"),
    _row("Kept you out of things on purpose etc from peer in the last 12 months"),
    _row("Past 12 months money or goods for sex", ["q507"]),
    _row("Most recent sexual touching in the last 12 months", ["q600", "q602"]),
    _row("Most recent attempted sex without consent/forced sex in the last 12 months", ["q700", "q702"]),
    _row("Most recent pressured into sex in the last 12 months", ["q900", "q902"]),
    _row("Most recent forced into sex in the last 12 months", ["q800", "q802"]),
    _row("Sex acts online in the last 12 months"),
    _row("Sent sexual photo/video online in the last 12 months"),
    _row("Anything else sexual online in the last 12 months"),
    _row("Sexually harassed online in the last 12 months"),
    _row("Hit etc your partner in the last 12 months", ["q200a"], male_only=True),
    _row("Hit with object etc your partner in the last 12 months"),
    _row("Choked etc your partner in the last 12 months", ["q200b"], male_only=True),
    _row("Threatened with knife/weapon etc your partner in the last 12 months", ["q200c"], male_only=True),
    _row("Hit etc someone else who is not your partner in the last 12 months", ["q201a"], male_only=True),
    _row("Hit with object etc someone else who is not your partner in the last 12 months"),
    _row("Choked etc someone else who is not your partner in the last 12 months", ["q201b"], male_only=True),
    _row("Threatened with knife/weapon etc someone else who is not your partner in the last 12 months", ["q201c"], male_only=True),
    _row("Forced your partner or ex to have sex in the last 12 months", ["q1100"], male_only=True),
    _row("Pressured your partner or ex to talk about sex online/virtual in the last 12 months"),
    _row("Pressured someone else who is not your partner to talk about sex online/virtual in the last 12 months"),
    _row("Pressured your partner or ex to send you sex material online/virtual in the last 12 months"),
    _row("Pressured someone else who is not your partner to send you sex material online/virtual in the last 12 months"),
    _row("Pressured your partner or ex to do anything else sexual online/virtual in the last 12 months"),
]


def _type_for(df: pd.DataFrame, vars_: list[str]) -> str:
    if not vars_:
        return ""
    first = vars_[0].lower()
    if first not in df.columns:
        return ""
    return classify_variable_type(df[first])


def main() -> None:
    codebook = _read_respondent_codebook()
    df, _ = pyreadstat.read_dta(PUD)
    rows: list[dict[str, str]] = []

    for spec in OUTCOME_SPEC:
        vars_ = list(spec["vars"])
        question_vars = list(spec["questions_vars"] or vars_[:1])
        if not vars_:
            rows.append({"Category": spec["category"], "Variable": "", "Type": "", "Format": "", "Questions": ""})
            continue

        if spec["male_only"]:
            questions = " || ".join(_male_question(v) for v in question_vars if _male_question(v))
        else:
            questions = _questions(codebook, question_vars)

        rows.append({
            "Category": spec["category"],
            "Variable": "; ".join(vars_),
            "Type": _type_for(df, vars_),
            "Format": _fmt(codebook, vars_),
            "Questions": questions,
        })

    out_df = build_outcome_codebook_df(rows)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_TSV.write_text(outcome_codebook_to_tsv(out_df), encoding="utf-8")
    export_minimal_codebook_docx(out_df, OUT_DOCX, columns=OUTCOME_CODEBOOK_COLUMNS, merge_category_column="Category")
    print(f"Wrote {OUT_DOCX}")
    print(f"Wrote {OUT_TSV}")
    print(f"Rows: {len(out_df)}")


if __name__ == "__main__":
    main()
