"""
Country-specific outcome question extraction helpers.

These helpers keep the orchestration in country scripts while sharing the
boring parts: PDF parsing, aligned Excel-codebook parsing, label-script parsing,
and TSV/DOCX export. They intentionally do not read respondent-level datasets.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import pandas as pd

from utils.docx_export import export_minimal_codebook_docx
from utils.pdf_parse import CodebookEntry, extract_codebook_entries, extract_questionnaire_questions


QUESTION_RESPONSE_COLUMNS: Sequence[str] = (
    "Country",
    "Instrument",
    "SourceType",
    "SourceFile",
    "QuestionNumber",
    "Variable",
    "Format",
    "Questions",
    "SourceStatus",
)


@dataclass(frozen=True)
class SourceSpec:
    """One explicit source file listed by a country script."""

    instrument: str
    path: Path
    source_type: str


def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _rel(path: Path) -> str:
    try:
        return str(path.relative_to(_project_root()))
    except ValueError:
        return str(path)


def _safe_format(entry: CodebookEntry) -> str:
    return entry.format_string.strip()


def _format_from_statistical_codebook_text(text: str) -> str:
    """Extract value labels from statistical codebook text blocks when parser cannot."""
    if "Values and labels" not in text:
        return ""
    # Curly apostrophes inside words (DON'T, child's) use the same glyph as
    # codebook quote delimiters in some PDFs; normalize them before matching.
    normalized = re.sub(r"(?<=\w)[’'‘](?=\w)", "", text)
    pairs = re.findall(r"\b(\d{1,4})\s+[’'‘\"]([^’'‘\"]+)[’'‘\"]", normalized)
    return ", ".join(f"{code}-{label.strip()}" for code, label in pairs)


def _clean_statistical_codebook_question(text: str) -> str:
    """Keep the quoted question stem from statistical codebook pages when available."""
    m = re.search(r'"\s*(.+?)"\s*-{5,}', text, flags=re.DOTALL)
    if not m:
        m = re.search(r'"""(.+?)"""', text, flags=re.DOTALL)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return text


def _row(
    *,
    country: str,
    source: SourceSpec,
    question_number: str = "",
    variable: str = "",
    fmt: str = "",
    question: str = "",
    source_status: str = "",
) -> dict[str, str]:
    return {
        "Country": country,
        "Instrument": source.instrument,
        "SourceType": source.source_type,
        "SourceFile": _rel(source.path),
        "QuestionNumber": question_number,
        "Variable": variable,
        "Format": fmt,
        "Questions": question,
        "SourceStatus": source_status,
    }


def parse_questionnaire_source(
    country: str,
    source: SourceSpec,
) -> list[dict[str, str]]:
    entries = extract_questionnaire_questions(source.path)
    return [
        _row(
            country=country,
            source=source,
            question_number=e.question_number,
            fmt=_safe_format(e),
            question=e.question_text,
        )
        for e in entries
    ]


def parse_codebook_source(
    country: str,
    source: SourceSpec,
) -> list[dict[str, str]]:
    entries = extract_codebook_entries(source.path)
    rows: list[dict[str, str]] = []
    for e in entries:
        if not (e.question_number or e.variable_name or e.question_text or e.response_options):
            continue
        variable = e.variable_name or e.question_number
        fmt = _safe_format(e) or _format_from_statistical_codebook_text(e.question_text)
        question = _clean_statistical_codebook_question(e.question_text)
        rows.append(
            _row(
                country=country,
                source=source,
                question_number=e.question_number,
                variable=variable,
                fmt=fmt,
                question=question,
            )
        )
    return rows


def _parse_xlsx_response_cell(value: object) -> tuple[str, str] | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"total", "frequency missing"}:
        return None
    m = re.match(r"^(\d{1,4})\s*=\s*(.+)$", text)
    if m:
        return m.group(1), m.group(2).strip()
    # Age/range rows often appear as bare values; keep them as self-labels.
    if re.match(r"^\d{1,4}(?:-\d{1,4})?$", text):
        return text, text
    return None


def parse_aligned_xlsx_source(
    country: str,
    source: SourceSpec,
) -> list[dict[str, str]]:
    """Parse aligned VACS Excel codebook blocks used by Namibia/Zimbabwe."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [
            _row(
                country=country,
                source=source,
                question="openpyxl is not installed; aligned xlsx source skipped.",
            )
        ]

    wb = load_workbook(source.path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    try:
        for ws in wb.worksheets:
            values = list(ws.iter_rows(values_only=True))
            for i, row_values in enumerate(values):
                first = "" if not row_values or row_values[0] is None else str(row_values[0]).strip()
                second = "" if len(row_values) < 2 or row_values[1] is None else str(row_values[1]).strip()
                if not first or second.lower() != "skip":
                    continue
                variable = first
                question = ""
                if i > 0 and values[i - 1] and values[i - 1][0] is not None:
                    question = str(values[i - 1][0]).strip()
                responses: list[tuple[str, str]] = []
                j = i + 1
                while j < len(values):
                    nxt_first = "" if not values[j] or values[j][0] is None else str(values[j][0]).strip()
                    nxt_second = "" if len(values[j]) < 2 or values[j][1] is None else str(values[j][1]).strip()
                    if not nxt_first:
                        break
                    if nxt_second.lower() == "skip":
                        break
                    parsed = _parse_xlsx_response_cell(nxt_first)
                    if parsed:
                        responses.append(parsed)
                    j += 1
                fmt = ", ".join(f"{code}-{label}" for code, label in responses)
                rows.append(
                    _row(
                        country=country,
                        source=source,
                        question_number=variable,
                        variable=variable,
                        fmt=fmt,
                        question=question,
                    )
                )
    finally:
        wb.close()
    return rows


def rows_to_df(rows: Iterable[Mapping[str, str]]) -> pd.DataFrame:
    return pd.DataFrame([{col: row.get(col, "") for col in QUESTION_RESPONSE_COLUMNS} for row in rows])


def export_question_response_table(df: pd.DataFrame, slug: str, *, output_dir: Path | None = None) -> None:
    out_dir = output_dir or (_project_root() / "data" / "processed" / "outcome")
    out_dir.mkdir(parents=True, exist_ok=True)
    tsv = out_dir / f"jcx_{slug}OutcomeQuestions.tsv"
    docx = out_dir / f"jcx_{slug}OutcomeQuestions.docx"
    df.to_csv(tsv, sep="\t", index=False)
    export_minimal_codebook_docx(df, docx, columns=QUESTION_RESPONSE_COLUMNS, merge_category_column=None)
    print(f"Wrote {docx}")
    print(f"Wrote {tsv}")
    print(f"Rows: {len(df)}")


def run_pdf_first_country(
    *,
    country: str,
    slug: str,
    questionnaire_sources: Sequence[SourceSpec],
    codebook_sources: Sequence[SourceSpec] = (),
    aligned_xlsx_sources: Sequence[SourceSpec] = (),
) -> pd.DataFrame:
    """Run one country's approved-document extraction and export outputs."""
    rows: list[dict[str, str]] = []
    for source in questionnaire_sources:
        rows.extend(parse_questionnaire_source(country, source))
    for source in codebook_sources:
        rows.extend(parse_codebook_source(country, source))
    for source in aligned_xlsx_sources:
        rows.extend(parse_aligned_xlsx_source(country, source))
    df = rows_to_df(rows)
    export_question_response_table(df, slug)
    return df


def rows_from_sas_or_do(country: str, source: SourceSpec) -> list[dict[str, str]]:
    """Extract conservative label-like statements from SAS/do files."""
    text = source.path.read_text(encoding="utf-8", errors="ignore")
    rows: list[dict[str, str]] = []
    patterns = [
        re.compile(r"^\s*label\s+([A-Za-z_]\w*)\s*=\s*['\"](.+?)['\"]", re.IGNORECASE),
        re.compile(r"^\s*label\s+variable\s+([A-Za-z_]\w*)\s+['\"](.+?)['\"]", re.IGNORECASE),
    ]
    for line in text.splitlines():
        for pat in patterns:
            m = pat.search(line)
            if m:
                rows.append(
                    _row(
                        country=country,
                        source=source,
                        question_number=m.group(1),
                        variable=m.group(1),
                        question=m.group(2),
                    )
                )
                break
    return rows


def run_label_script_country(
    *,
    country: str,
    slug: str,
    script_sources: Sequence[SourceSpec],
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for source in script_sources:
        rows.extend(rows_from_sas_or_do(country, source))
    df = rows_to_df(rows)
    export_question_response_table(df, slug)
    if not len(df):
        print("Warning: no question/response metadata rows were extracted.", file=sys.stderr)
    return df
